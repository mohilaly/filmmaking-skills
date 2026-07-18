#!/usr/bin/env python3
"""Build the Excel workbook from weekly_data.jsonl + films_data.jsonl.

Films are aggregated by canonical film_id (config/films_canonical.json), so a
film that appears under different title spellings counts as one film.
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import bo_lib
from bo_lib import ROOT, PAGE

OUT = ROOT / "Saudi_Box_Office_Weekly.xlsx"

# --- Styles ---
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")


def style_header_row(ws, n_cols):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def style_zebra(ws, n_cols):
    for row in ws.iter_rows(min_row=2, max_col=n_cols):
        for c in row:
            c.border = BORDER
        if row[0].row % 2 == 0:
            for c in row:
                c.fill = ALT_FILL


def main():
    weeks = bo_lib.load_weeks()
    films_by_week = bo_lib.load_films_raw()
    canon = bo_lib.load_canon()

    weeks.sort(key=lambda r: r.get("date_end") or "", reverse=True)

    wb = Workbook()

    # ---------- Weekly Summary ----------
    ws = wb.active
    ws.title = "Weekly Summary"
    headers = [
        "Week ending (Sun)", "Year", "Month", "Week #",
        "Films in cinema", "Tickets sold (thousands)",
        "Revenue (million SAR)", "Avg ticket price (SAR)",
        "Week label (AR)", "Date start", "Date end",
        "Source image", "Image URL",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))
    for r in weeks:
        avg = None
        if r.get("revenue_M_SAR") and r.get("tickets_K"):
            avg = round((r["revenue_M_SAR"] * 1_000_000) / (r["tickets_K"] * 1_000), 2)
        ws.append([
            r.get("date_end") or "", r.get("year") or "", r.get("month") or "",
            r.get("week") or "",
            r.get("films"), r.get("tickets_K"), r.get("revenue_M_SAR"), avg,
            r.get("week_label_ar") or "",
            r.get("date_start") or "", r.get("date_end") or "",
            r["filename"], bo_lib.img_url(r["filename"]),
        ])
    for i, w in enumerate([16, 7, 7, 7, 13, 16, 17, 13, 26, 12, 12, 50, 70], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    style_zebra(ws, len(headers))
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for c in row:
            if c.column == 9:
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif c.column in (12, 13):
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    # ---------- Top 10 Films per Week ----------
    ws_films = wb.create_sheet("Top 10 Films")
    cols = [
        "Week ending", "Year", "Month", "Rank",
        "Title (Arabic)", "Title (English)", "Country",
        "Weeks in cinema", "Week revenue (M SAR)", "Week tickets (K)",
        "Total revenue (M SAR)", "Total tickets (K)",
        "Film ID", "Source image",
    ]
    ws_films.append(cols)
    style_header_row(ws_films, len(cols))

    rows = bo_lib.film_rows(weeks, films_by_week, canon)
    for r in rows:
        ws_films.append([
            r["week_end"], r["year"], r["month"], r["rank"],
            r["title_ar"], r["title_en"], r["country"],
            r["weeks_in_cinema"], r["week_revenue_M"], r["week_tickets_K"],
            r["total_revenue_M"], r["total_tickets_K"],
            r["film_id"] or "?", r["filename"],
        ])
    for i, w in enumerate([14, 7, 7, 7, 26, 32, 13, 9, 14, 14, 14, 14, 22, 50], 1):
        ws_films.column_dimensions[get_column_letter(i)].width = w
    style_zebra(ws_films, len(cols))
    for row in ws_films.iter_rows(min_row=2, max_col=len(cols)):
        for c in row:
            if c.column == 5:  # arabic title
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif c.column in (6, 13, 14):
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    # ---------- Films by Title (keyed on canonical film_id) ----------
    ws_pivot = wb.create_sheet("Films by Title")
    cols2 = [
        "Title (English)", "Title (Arabic)", "Country", "Genre", "First week", "Last week",
        "Weeks in top 10", "Best rank",
        "Best week revenue (M)", "Best total revenue (M)", "Best total tickets (K)",
        "Film ID",
    ]
    ws_pivot.append(cols2)
    style_header_row(ws_pivot, len(cols2))

    by_film = {}
    for r in rows:
        fid = r["film_id"]
        if fid is None:
            continue  # unresolved rows are caught by integrity check H
        d = by_film.setdefault(fid, {
            "weeks": [], "ranks": [],
            "best_week_rev": 0.0, "best_total_rev": 0.0, "best_total_tk": 0.0,
        })
        d["weeks"].append(r["week_end"])
        if r["rank"]:
            d["ranks"].append(r["rank"])
        if r["week_revenue_M"] and r["week_revenue_M"] > d["best_week_rev"]:
            d["best_week_rev"] = r["week_revenue_M"]
        if r["total_revenue_M"] and r["total_revenue_M"] > d["best_total_rev"]:
            d["best_total_rev"] = r["total_revenue_M"]
        if r["total_tickets_K"] and r["total_tickets_K"] > d["best_total_tk"]:
            d["best_total_tk"] = r["total_tickets_K"]

    title_rows = []
    for fid, d in by_film.items():
        cf = canon.get(fid)
        weeks_sorted = sorted(w for w in d["weeks"] if w)
        title_rows.append({
            "en": cf.get("title_en") or "", "ar": cf.get("title_ar") or "",
            "country": cf.get("country") or "", "genre": cf.get("genre") or "",
            "first": weeks_sorted[0] if weeks_sorted else "",
            "last": weeks_sorted[-1] if weeks_sorted else "",
            "n_weeks": len(d["weeks"]),
            "best_rank": min(d["ranks"]) if d["ranks"] else "",
            "best_wrev": round(d["best_week_rev"], 3) if d["best_week_rev"] else None,
            "best_trev": round(d["best_total_rev"], 3) if d["best_total_rev"] else None,
            "best_ttk": round(d["best_total_tk"], 1) if d["best_total_tk"] else None,
            "film_id": fid,
        })
    title_rows.sort(key=lambda r: (r["best_trev"] or 0), reverse=True)

    for r in title_rows:
        ws_pivot.append([
            r["en"] or r["ar"], r["ar"], r["country"], r["genre"], r["first"], r["last"],
            r["n_weeks"], r["best_rank"],
            r["best_wrev"], r["best_trev"], r["best_ttk"], r["film_id"],
        ])
    for i, w in enumerate([34, 26, 13, 16, 13, 13, 14, 11, 18, 19, 19, 22], 1):
        ws_pivot.column_dimensions[get_column_letter(i)].width = w
    style_zebra(ws_pivot, len(cols2))
    for row in ws_pivot.iter_rows(min_row=2, max_col=len(cols2)):
        for c in row:
            if c.column == 2:  # arabic
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif c.column in (1, 12):
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")

    # ---------- Monthly Totals ----------
    ws2 = wb.create_sheet("Monthly Totals")
    ws2.append(["Year", "Month", "Weeks", "Total tickets (K)", "Total revenue (M SAR)", "Avg films/week", "Avg ticket price"])
    style_header_row(ws2, 7)
    by_month = {}
    for r in weeks:
        if not r.get("year") or not r.get("month"):
            continue
        key = (r["year"], r["month"])
        d = by_month.setdefault(key, {"weeks": 0, "tk": 0.0, "rv": 0.0, "films": []})
        d["weeks"] += 1
        if r.get("tickets_K"): d["tk"] += r["tickets_K"]
        if r.get("revenue_M_SAR"): d["rv"] += r["revenue_M_SAR"]
        if r.get("films"): d["films"].append(r["films"])
    for key in sorted(by_month, reverse=True):
        d = by_month[key]
        avg_films = round(sum(d["films"]) / len(d["films"]), 1) if d["films"] else None
        avg_price = round(d["rv"] * 1_000_000 / (d["tk"] * 1_000), 2) if d["tk"] else None
        ws2.append([key[0], key[1], d["weeks"], round(d["tk"], 1), round(d["rv"], 1), avg_films, avg_price])
    for i, w in enumerate([8, 8, 8, 18, 22, 16, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    style_zebra(ws2, 7)
    for row in ws2.iter_rows(min_row=2, max_col=7):
        for c in row:
            c.alignment = Alignment(horizontal="center")

    # ---------- Yearly Totals ----------
    ws3 = wb.create_sheet("Yearly Totals")
    ws3.append(["Year", "Weeks captured", "Total tickets (K)", "Total revenue (M SAR)", "Avg films/week", "Avg ticket price (SAR)"])
    style_header_row(ws3, 6)
    by_year = {}
    for r in weeks:
        if not r.get("year"): continue
        d = by_year.setdefault(r["year"], {"weeks": 0, "tk": 0.0, "rv": 0.0, "films": []})
        d["weeks"] += 1
        if r.get("tickets_K"): d["tk"] += r["tickets_K"]
        if r.get("revenue_M_SAR"): d["rv"] += r["revenue_M_SAR"]
        if r.get("films"): d["films"].append(r["films"])
    for y in sorted(by_year, reverse=True):
        d = by_year[y]
        avg_films = round(sum(d["films"]) / len(d["films"]), 1) if d["films"] else None
        avg_price = round(d["rv"] * 1_000_000 / (d["tk"] * 1_000), 2) if d["tk"] else None
        ws3.append([y, d["weeks"], round(d["tk"], 1), round(d["rv"], 1), avg_films, avg_price])
    for i, w in enumerate([8, 16, 18, 22, 16, 22], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    style_zebra(ws3, 6)
    for row in ws3.iter_rows(min_row=2, max_col=6):
        for c in row:
            c.alignment = Alignment(horizontal="center")

    # ---------- About ----------
    ws4 = wb.create_sheet("About")
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 90
    n_gaps = len(bo_lib.known_gaps(weeks))
    info = [
        ("Source", PAGE),
        ("Publisher", "Saudi Film Commission (هيئة الأفلام) — Ministry of Culture"),
        ("Coverage", f"{weeks[-1].get('date_end','?')}  →  {weeks[0].get('date_end','?')} ({len(weeks)} weeks; {n_gaps} known missing weeks — see Coverage Calendar)"),
        ("Films extracted", f"{sum(len(f.get('films',[])) for f in films_by_week)} film entries across {len(films_by_week)} weeks; {len(title_rows)} unique films (canonical registry)"),
        ("Units", "tickets in thousands; revenue in millions of Saudi Riyals (SAR)"),
        ("Avg ticket price", "Computed as revenue × 1,000,000 ÷ (tickets × 1,000)"),
        ("Film identity", "Films are unified via config/films_canonical.json — the same film under different title spellings counts once."),
        ("Method", "Each weekly report image was downloaded, cropped to its header and to two film-row strips, and read with vision to extract: films-in-cinema/tickets/revenue summary plus the top-10 films panel (rank, title, country, week revenue, total revenue, week tickets, total tickets)."),
        ("Notes", "Reports cover the prior Sun-Sat trading week (date in column 'Week ending' is the report's published date). Some movie titles in early 2024 reports show only integer values; from late 2024 onward decimals are included. Cumulative ticket figures show total across all weeks since film release."),
        ("Generated", "Auto-extracted; spot-check against original images before quoting publicly."),
    ]
    for i, (k, v) in enumerate(info, 1):
        a = ws4.cell(row=i, column=1, value=k)
        a.font = Font(bold=True)
        a.alignment = Alignment(vertical="top")
        b = ws4.cell(row=i, column=2, value=v)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        ws4.row_dimensions[i].height = 30 if len(v) > 80 else 18

    wb._sheets = [ws, ws_films, ws_pivot, ws2, ws3, ws4]

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")
    print(f"  Weekly Summary: {len(weeks)} rows")
    print(f"  Top 10 Films: {len(rows)} rows")
    print(f"  Films by Title: {len(title_rows)} unique films")
    unresolved = [r for r in rows if r["film_id"] is None]
    if unresolved:
        print(f"  WARNING: {len(unresolved)} film rows did not resolve to a canonical film")


if __name__ == "__main__":
    main()
