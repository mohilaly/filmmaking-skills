#!/usr/bin/env python3
"""Build the 'Coverage Calendar' sheet — text table + visual heatmap side-by-side.

Left side (cols A–F): a detailed table — each year, months down × W1–W5 across,
with the Saturday date and a status marker (✓ / MISSING / —).

Right side (cols H–BJ): a compact visual heatmap — one row per year, one tiny
colored cell per Saturday of the year (53 cells per row max). No text, just colors,
so you can see coverage gaps at a glance.

Color coding (both views):
  green  = data captured for that week
  red    = expected coverage but missing
  grey   = before publishing started or future
"""
from __future__ import annotations
import json
import datetime
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC_WEEKS = ROOT / "data" / "weekly_data.jsonl"
OUT_XLSX = ROOT / "Saudi_Box_Office_Weekly.xlsx"

def _data_bounds():
    dates = []
    for line in SRC_WEEKS.open(encoding="utf-8"):
        r = json.loads(line)
        if r.get("date_end"):
            dates.append(datetime.date.fromisoformat(r["date_end"]))
    return min(dates), max(dates)

_FIRST, _LAST = _data_bounds()
COVERAGE_START_SAT = _FIRST
TODAY = _LAST  # cutoff = last captured week; later Saturdays are "not yet", not missing
YEARS = list(range(_FIRST.year - 1, TODAY.year + 1))
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def week_label_month(saturday: datetime.date) -> tuple[int, int]:
    counts: dict[tuple[int, int], int] = defaultdict(int)
    for i in range(7):
        d = saturday - datetime.timedelta(days=i)
        counts[(d.year, d.month)] += 1
    return max(counts, key=lambda k: (counts[k], k))


def all_saturdays(start: datetime.date, end: datetime.date):
    d = start + datetime.timedelta(days=(5 - start.weekday()) % 7)
    while d <= end:
        yield d
        d += datetime.timedelta(days=7)


def index_records():
    by_date = {}
    for line in SRC_WEEKS.open(encoding="utf-8"):
        r = json.loads(line)
        de = r.get("date_end")
        if de:
            by_date[de] = r
    return by_date


def saturday_status(sat: datetime.date, records_by_date: dict) -> str:
    """Return one of: 'have', 'missing', 'outside'."""
    if sat < COVERAGE_START_SAT or sat > TODAY:
        return "outside"
    return "have" if sat.isoformat() in records_by_date else "missing"


def main():
    records_by_date = index_records()

    period_start = datetime.date(2023, 1, 1)
    period_end = datetime.date(2026, 12, 31)
    saturdays = list(all_saturdays(period_start, period_end))

    by_year_month: dict[tuple[int, int], list[datetime.date]] = defaultdict(list)
    for sat in saturdays:
        ym = week_label_month(sat)
        by_year_month[ym].append(sat)

    # Open existing workbook
    wb = load_workbook(OUT_XLSX)
    if "Coverage Calendar" in wb.sheetnames:
        del wb["Coverage Calendar"]
    ws = wb.create_sheet("Coverage Calendar")

    # Styles
    title_font = Font(bold=True, size=14, color="1F4E78")
    year_font = Font(bold=True, size=12, color="FFFFFF")
    year_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    have_fill = PatternFill("solid", fgColor="63BE7B")  # bright green
    have_font = Font(bold=True, color="0E5C2F", size=9)
    miss_fill = PatternFill("solid", fgColor="F8696B")  # bright red
    miss_font = Font(bold=True, color="9C0006", size=9)
    outside_fill = PatternFill("solid", fgColor="E7E6E6")  # light grey
    outside_font = Font(color="999999", italic=True, size=9)
    none_fill = PatternFill("solid", fgColor="FAFAFA")
    thin = Side(border_style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    thin_white = Side(border_style="thin", color="FFFFFF")
    cell_border = Border(left=thin_white, right=thin_white, top=thin_white, bottom=thin_white)

    # ---------- Title and legend ----------
    ws["A1"] = "Saudi Box Office — Weekly Coverage Calendar"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")

    ws["A2"] = (
        "Each cell is one Sun–Sat trading week. "
        "Green ✓ = captured · Red MISSING = expected but absent · Grey — = before reports began / future. "
        f"Coverage starts {COVERAGE_START_SAT.strftime('%a %d %b %Y')}; cutoff {TODAY.strftime('%a %d %b %Y')}."
    )
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 35

    # Stats
    n_have = n_missing = n_outside = 0
    for sat in saturdays:
        if sat.year not in YEARS:
            continue
        s = saturday_status(sat, records_by_date)
        if s == "have": n_have += 1
        elif s == "missing": n_missing += 1
        else: n_outside += 1

    ws["A3"] = f"Captured: {n_have} weeks   ·   Missing: {n_missing} weeks   ·   Outside coverage window: {n_outside} weeks"
    ws["A3"].font = Font(bold=True, size=10)
    ws.merge_cells("A3:F3")

    # ---------- LEFT: detailed table ----------
    row = 5
    for c, w in {1: 16, 2: 12, 3: 12, 4: 12, 5: 12, 6: 12}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    for year in YEARS:
        ws.cell(row=row, column=1, value=str(year))
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = year_fill
            cell.font = year_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 22
        row += 1

        ws.cell(row=row, column=1, value="Month").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = border
        for i in range(1, 6):
            c = ws.cell(row=row, column=1 + i, value=f"W{i}")
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            c.border = border
        row += 1

        for m in range(1, 13):
            ws.cell(row=row, column=1, value=MONTHS[m - 1]).font = Font(bold=True)
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
            sats_in_month = by_year_month.get((year, m), [])
            for w_idx in range(5):
                cell = ws.cell(row=row, column=2 + w_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if w_idx < len(sats_in_month):
                    sat = sats_in_month[w_idx]
                    s = saturday_status(sat, records_by_date)
                    if s == "outside":
                        cell.value = sat.strftime("%d %b") + "\n—"
                        cell.fill = outside_fill
                        cell.font = outside_font
                    elif s == "have":
                        cell.value = sat.strftime("%d %b") + "\n✓"
                        cell.fill = have_fill
                        cell.font = have_font
                    else:
                        cell.value = sat.strftime("%d %b") + "\nMISSING"
                        cell.fill = miss_fill
                        cell.font = miss_font
                else:
                    cell.fill = none_fill
            ws.row_dimensions[row].height = 28
            row += 1
        row += 1  # spacer between years

    # ---------- RIGHT: visual heatmap ----------
    # Layout: starting at col H (col 8). One row per year, one cell per Saturday.
    # Use month markers above to label timeline.

    HEAT_FIRST_COL = 8  # column H

    # Heatmap title
    ws.cell(row=5, column=HEAT_FIRST_COL, value="Visual Coverage Timeline").font = title_font
    ws.merge_cells(start_row=5, start_column=HEAT_FIRST_COL, end_row=5, end_column=HEAT_FIRST_COL + 53)

    ws.cell(row=6, column=HEAT_FIRST_COL, value=(
        "One small cell per week of each year. Hover/click to see the date. "
        "Use this to scan for gaps at a glance."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=6, start_column=HEAT_FIRST_COL, end_row=6, end_column=HEAT_FIRST_COL + 53)
    ws.row_dimensions[6].height = 30

    # For each year, lay out 1 row of Saturdays
    # Column widths for heatmap cells - tight
    for c in range(HEAT_FIRST_COL, HEAT_FIRST_COL + 54):
        ws.column_dimensions[get_column_letter(c)].width = 3.0

    # Year label column to the left of the heatmap row needs more width
    ws.column_dimensions[get_column_letter(HEAT_FIRST_COL)].width = 7

    heatmap_start_row = 8

    # Month markers row — show month abbrevs above each year's strip
    # For a given year, the Saturdays of that year are heatmap cells 1..N (N=52 or 53).
    # Month markers will be drawn above each year's row.

    cur_row = heatmap_start_row
    for year in YEARS:
        # Year saturdays in chronological order
        year_sats = [s for s in saturdays if s.year == year]
        # Month marker line above year's strip
        # Build month labels: place month name at the column corresponding to first Saturday of that month
        first_sat_col_by_month = {}
        for idx, sat in enumerate(year_sats):
            if sat.month not in first_sat_col_by_month:
                first_sat_col_by_month[sat.month] = HEAT_FIRST_COL + 1 + idx

        for m in range(1, 13):
            if m in first_sat_col_by_month:
                cell = ws.cell(row=cur_row, column=first_sat_col_by_month[m], value=MONTHS[m - 1])
                cell.font = Font(bold=True, size=8, color="555555")
                cell.alignment = Alignment(horizontal="left", vertical="bottom")
        ws.row_dimensions[cur_row].height = 14
        cur_row += 1

        # Year strip: label + 52/53 colored cells
        yc = ws.cell(row=cur_row, column=HEAT_FIRST_COL, value=str(year))
        yc.font = Font(bold=True, size=11, color="1F4E78")
        yc.alignment = Alignment(horizontal="right", vertical="center")

        for idx, sat in enumerate(year_sats):
            col = HEAT_FIRST_COL + 1 + idx
            cell = ws.cell(row=cur_row, column=col)
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            s = saturday_status(sat, records_by_date)
            if s == "have":
                cell.fill = have_fill
            elif s == "missing":
                cell.fill = miss_fill
            else:
                cell.fill = outside_fill
            # Tooltip-style: store the date in the cell value as a small note (Excel doesn't have hover natively here)
            cell.value = sat.strftime("%d/%m")
            cell.font = Font(size=6, color="555555")

        ws.row_dimensions[cur_row].height = 28
        cur_row += 2  # gap between years

    # Heatmap legend below the strip
    legend_row = cur_row + 1
    ws.cell(row=legend_row, column=HEAT_FIRST_COL, value="Legend:").font = Font(bold=True)
    legend_items = [
        ("Captured", have_fill, have_font),
        ("Missing", miss_fill, miss_font),
        ("Outside coverage", outside_fill, outside_font),
    ]
    col = HEAT_FIRST_COL + 2
    for label, fill, font in legend_items:
        sq = ws.cell(row=legend_row, column=col, value=" ")
        sq.fill = fill
        sq.border = cell_border
        ws.cell(row=legend_row, column=col + 1, value=label).font = Font(size=10)
        col += 4

    ws.freeze_panes = "A5"

    # Reorder sheets: Weekly Summary, Coverage Calendar, Top 10 Films, ...
    sheets = list(wb._sheets)
    cal = sheets.pop()  # we just added it last
    sheets.insert(1, cal)
    wb._sheets = sheets

    wb.save(OUT_XLSX)
    print(f"Updated {OUT_XLSX}")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()
